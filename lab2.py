import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook

class TabbedApp:
    def __init__(self, root):
        self.root = root
        root.minsize(640, 480)
        root.maxsize(640, 480)
        self.root.title("Проектирование склада заполнителей (Площадь)")
        self.root.resizable(False, False)
        self.vcmd = (root.register(self._validate_volume), '%P')
        self.top_frame = tk.Frame(root)
        self.top_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        self.warehouse_type_label = tk.Label(self.top_frame, text="Тип склада:")
        self.warehouse_type_label.grid(row=0, column=0, sticky=tk.W)
        self.warehouse_type_var = tk.StringVar(value="Штабельный")
        self.warehouse_type_combobox = ttk.Combobox(self.top_frame, textvariable=self.warehouse_type_var, state="readonly")
        self.warehouse_type_combobox['values'] = ("Штабельный", "Другой")
        self.warehouse_type_combobox.grid(row=0, column=1, sticky=tk.W)
        self.warehouse_type_combobox.bind("<<ComboboxSelected>>", self.on_warehouse_type_change)
        self.usage_coefficient_label = tk.Label(self.top_frame, text="Кис:")
        self.usage_coefficient_label.grid(row=1, column=0, sticky=tk.W, pady=(10, 0))
        self.usage_coefficient_var = tk.DoubleVar(value=0.7)
        self.usage_coefficient_spinbox = tk.Spinbox(
            self.top_frame,
            from_=0.7,
            to=0.8,
            increment=0.01,
            textvariable=self.usage_coefficient_var,
            format="%.2f",
            width=5,
            state="readonly"
        )
        self.usage_coefficient_spinbox.grid(row=1, column=1, sticky=tk.W, pady=(10, 0))
        self.top_frame.grid_columnconfigure(1, weight=1)
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(expand=True, fill='both')
        self.create_tab("Материал 1")

        self.buttons_frame = tk.Frame(root)
        self.buttons_frame.pack(fill=tk.X, padx=10, pady=(5, 0))

        self.add_tab_button = tk.Button(self.buttons_frame, text="Добавить материал", command=self.add_tab)
        self.add_tab_button.grid(row=0, column=0, sticky="w")

        self.calculate_button = tk.Button(self.buttons_frame, text="Рассчитать", command=self.calculate_areas)
        self.calculate_button.grid(row=0, column=0, sticky="w", padx=125)

        self.export_button = tk.Button(self.buttons_frame, text="Экспортировать", command=self.export_to_excel)
        self.export_button.grid(row=0, column=0, sticky="w", padx=203)

        self.area_output_label = tk.Label(self.buttons_frame, text="Площади материалов (м²):")
        self.area_output_label.grid(row=1, column=0, columnspan=2, sticky="w", pady=(10, 0))
        self.area_output_text = tk.Text(self.buttons_frame, height=5, width=40, state='disabled')
        self.area_output_text.grid(row=2, column=0, columnspan=2, sticky="we", pady=(0, 10))

        self.buttons_frame.grid_columnconfigure(0, weight=1)
        self.buttons_frame.grid_columnconfigure(1, weight=1)

    def calculate_areas(self):
        self.update_all_areas()

    def export_to_excel(self):

        from openpyxl.styles import Font, Alignment, Border, Side

        export_file = "Проектирование склада заполнителей.xlsx"

        wb = Workbook()
        ws = wb.active

        headers = {
            "A1": "Наименование материала",
            "B1": "Vс.з.",
            "C1": "q",
            "D1": "Sм",
            "F1": "Тип склада",
            "G1": "Кис",
            "H1": "Sобщ"
        }
        for cell, value in headers.items():
            ws[cell] = value
            ws[cell].font = Font(bold=True)
            ws[cell].alignment = Alignment(horizontal="center", vertical="center")

        column_widths = {
            "A": 25,
            "B": 12,
            "C": 8,
            "D": 12,
            "F": 15,
            "G": 8,
            "H": 12
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        ws["F2"] = self.warehouse_type_var.get()
        ws["G2"] = self.usage_coefficient_var.get()

        total_area = 0.0
        for i in range(self.notebook.index("end")):
            tab = self.notebook.nametowidget(self.notebook.tabs()[i])
            area = self.update_area(tab)
            if area is not None:
                total_area += area
        ws["H2"] = total_area

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 2
        for i in range(self.notebook.index("end")):
            tab = self.notebook.nametowidget(self.notebook.tabs()[i])
            material_name = tab.material_name_entry.get()
            volume = tab.volume_var.get()
            q = tab.material_per_area_var.get()
            area = self.update_area(tab)
            ws.cell(row=row, column=1, value=material_name)
            ws.cell(row=row, column=2, value=volume)
            ws.cell(row=row, column=3, value=q)
            if area is not None:
                ws.cell(row=row, column=4, value=round(area, 2))
            else:
                ws.cell(row=row, column=4, value=None)

            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1

        for col, cell_ref in zip(["F", "G", "H"], ["F2", "G2", "H2"]):
            cell = ws[cell_ref]
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(export_file)

    def _validate_volume(self, new_value):
        if new_value == "":
            return True
        try:
            val = float(new_value)
            if val > 0 and val <= 1000:
                return True
            else:
                return False
        except ValueError:
            return False

    def create_tab(self, name):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=name)
        
        tk.Label(tab, text="Наименование материала:").grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        tab.material_name_entry = tk.Entry(tab)
        tab.material_name_entry.grid(row=0, column=1, padx=10, pady=5, sticky=tk.EW)

        close_button = tk.Button(tab, text="✕", command=lambda: self.close_tab(name))
        close_button.grid(row=0, column=2, padx=10, pady=5, sticky=tk.NE)

        tk.Label(tab, text="Объем склада заполнителей (Vс.з.), м³:").grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        tab.volume_var = tk.StringVar()
        tab.volume_entry = tk.Entry(tab, textvariable=tab.volume_var, validate='key', validatecommand=self.vcmd)
        tab.volume_entry.grid(row=1, column=1, padx=10, pady=5, sticky=tk.EW)

        tk.Label(tab, text="Количество материала на 1 м² (q), м³/м²:").grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
        tab.material_per_area_var = tk.DoubleVar()
        tab.material_per_area_spinbox = tk.Spinbox(
            tab,
            from_=3.0,
            to=4.0,
            increment=0.1,
            textvariable=tab.material_per_area_var,
            format="%.1f",
            width=5,
            state="readonly"
        )
        tab.material_per_area_spinbox.grid(row=2, column=1, padx=10, pady=5, sticky=tk.W)

        tab.volume_var.trace_add('write', lambda *args, tab=tab: self.update_all_areas())
        tab.material_per_area_var.trace_add('write', lambda *args, tab=tab: self.update_all_areas())
        self.usage_coefficient_var.trace_add('write', lambda *args: self.update_all_areas())

        close_button = tk.Button(tab, text="✕", command=lambda: self.close_tab(name))
        close_button.grid(row=0, column=2, padx=10, pady=5, sticky=tk.NE)

        tk.Label(tab, text="Объем склада заполнителей (Vс.з.), м³:").grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        tab.volume_var = tk.StringVar()
        tab.volume_entry = tk.Entry(tab, textvariable=tab.volume_var, validate='key', validatecommand=self.vcmd)
        tab.volume_entry.grid(row=1, column=1, padx=10, pady=5, sticky=tk.EW)

        tk.Label(tab, text="Количество материала на 1 м² (q), м³/м²:").grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
        tab.material_per_area_var = tk.DoubleVar()
        tab.material_per_area_spinbox = tk.Spinbox(
            tab,
            from_=3.0,
            to=4.0,
            increment=0.1,
            textvariable=tab.material_per_area_var,
            format="%.1f",
            width=5,
            state="readonly"
        )
        tab.material_per_area_spinbox.grid(row=2, column=1, padx=10, pady=5, sticky=tk.W)

        tab.grid_columnconfigure(1, weight=1)

    def add_tab(self):
        existing_tabs = [self.notebook.tab(i, "text") for i in range(self.notebook.index("end"))]
        new_index = len(existing_tabs) + 1
        new_tab_name = f"Материал {new_index}"
        
        self.create_tab(new_tab_name)
        self.update_material_per_area_spinboxes()

        self.notebook.select(self.notebook.index("end") - 1)

    def close_tab(self, name):
        for index in range(self.notebook.index("end")):
            if self.notebook.tab(index, "text") == name:
                self.notebook.forget(index)
                break

    def on_warehouse_type_change(self, event):
        self.update_material_per_area_spinboxes()

    def update_area(self, tab):
        try:
            volume = float(tab.volume_var.get())
            material_per_area = tab.material_per_area_var.get()
            usage_coefficient = self.usage_coefficient_var.get()
            if material_per_area > 0 and usage_coefficient > 0:
                area = volume / (material_per_area / usage_coefficient)
                return area
            else:
                return None
        except (ValueError, ZeroDivisionError):
            return None

    def update_all_areas(self):
        areas = []
        total_area = 0.0
        for i in range(self.notebook.index("end")):
            tab = self.notebook.nametowidget(self.notebook.tabs()[i])
            area = self.update_area(tab)
            if area is not None:
                areas.append(f"S{i+1} = {area:.2f} м²")
                total_area += area
        areas.append(f"Sобщ = {total_area:.2f} м²")
        self.area_output_text.config(state='normal')
        self.area_output_text.delete('1.0', tk.END)
        self.area_output_text.insert(tk.END, "\n".join(areas))
        self.area_output_text.config(state='disabled')

    def update_material_per_area_spinboxes(self):
        warehouse_type = self.warehouse_type_var.get()
        if warehouse_type == "Штабельный":
            from_val = 3.0
            to_val = 4.0
        else:
            from_val = 5.0
            to_val = 7.0

        for i in range(self.notebook.index("end")):
            tab = self.notebook.nametowidget(self.notebook.tabs()[i])
            if hasattr(tab, "material_per_area_spinbox"):
                spinbox = tab.material_per_area_spinbox
                var = tab.material_per_area_var
                spinbox.config(from_=from_val, to=to_val)
                current_val = var.get()
                if current_val < from_val or current_val > to_val:
                    var.set(from_val)

root = tk.Tk()
app = TabbedApp(root)
root.mainloop()